"""Excel 出力共通サービス — openpyxl ベース。

見積内訳書 / 出来高調書を A4 印刷想定の xlsx として bytes 化する。
例外は ValueError で投げ、呼び出し側 (ルータ層) で HTTPException 等に変換する。
"""

from decimal import Decimal
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

from models.estimate_section import EstimateSection
from models.estimate_item import EstimateItem
from models.progress_statement import ProgressStatement
from models.progress_row import ProgressRow
from models.progress_entry import ProgressEntry


# ---- スタイル定数 ----
_THIN = Side(border_style="thin", color="888888")
_MEDIUM = Side(border_style="medium", color="1a1a1a")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_HEADER = Border(left=_THIN, right=_THIN, top=_MEDIUM, bottom=_MEDIUM)

_HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", start_color="1A1A1A", end_color="1A1A1A")
_TITLE_FONT = Font(name="Yu Gothic", size=18, bold=True, color="1A1A1A")
_SUBTITLE_FONT = Font(name="Yu Gothic", size=11, color="555555")
_BODY_FONT = Font(name="Yu Gothic", size=10, color="1A1A1A")
_SUBTOTAL_FONT = Font(name="Yu Gothic", size=10, bold=True, color="1A1A1A")
_SUBTOTAL_FILL = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
_TOTAL_FILL = PatternFill(fill_type="solid", start_color="E8E8E8", end_color="E8E8E8")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _to_float(value) -> float:
    """Decimal / int / None を float に正規化。"""
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def _to_int(value) -> int:
    """Decimal / int / None を int に正規化。"""
    if value is None:
        return 0
    if isinstance(value, Decimal):
        return int(value)
    return int(value)


def _setup_a4_landscape(ws: Worksheet) -> None:
    """A4 横向き印刷想定の共通ページ設定。"""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6)
    ws.print_options.horizontalCentered = True


def _setup_a4_portrait(ws: Worksheet) -> None:
    """A4 縦向き印刷想定の共通ページ設定 (中表紙用)。"""
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=1.0, bottom=1.0)
    ws.print_options.horizontalCentered = True


def _write_cover_sheet(ws: Worksheet, estimate) -> None:
    """中表紙シート。タイトル + 案件名 + 顧客名 + 合計金額。"""
    _setup_a4_portrait(ws)
    # 列幅
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.merge_cells("A2:F2")
    cell = ws["A2"]
    cell.value = "工 事 費 内 訳 書"
    cell.font = Font(name="Yu Gothic", size=26, bold=True, color="1A1A1A")
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[2].height = 50

    ws.merge_cells("A4:F4")
    sub = ws["A4"]
    sub.value = "ESTIMATE BREAKDOWN"
    sub.font = _SUBTITLE_FONT
    sub.alignment = _ALIGN_CENTER

    rows_info = [
        ("案 件 名", getattr(estimate, "project_name", "") or ""),
        ("見 積 番 号", getattr(estimate, "estimate_number", "") or ""),
        ("顧 客 名", getattr(estimate, "customer_name", "") or ""),
        ("合 計 金 額 (税込)", f"¥{_to_int(getattr(estimate, 'total', 0)):,}"),
    ]
    row_idx = 8
    for label, value in rows_info:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = _SUBTOTAL_FONT
        label_cell.fill = _SUBTOTAL_FILL
        label_cell.alignment = _ALIGN_CENTER
        label_cell.border = _BORDER_ALL

        ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
        val_cell = ws.cell(row=row_idx, column=3, value=value)
        val_cell.font = _BODY_FONT
        val_cell.alignment = _ALIGN_LEFT
        val_cell.border = _BORDER_ALL
        ws.row_dimensions[row_idx].height = 28
        row_idx += 2


def _write_detail_sheet(
    ws: Worksheet,
    estimate,
    sections_with_items: Iterable[tuple[EstimateSection, list[EstimateItem]]],
) -> None:
    """内訳書本体シート。"""
    _setup_a4_landscape(ws)
    headers = ["No", "種別 / 項目", "仕様", "数量", "単位", "単価", "金額", "備考"]
    widths = [5, 28, 24, 10, 8, 14, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- タイトル行 ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="工 事 費 内 訳 書")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    # ---- 案件メタ ----
    meta_row = 3
    project_name = getattr(estimate, "project_name", "") or ""
    customer_name = getattr(estimate, "customer_name", "") or ""
    estimate_number = getattr(estimate, "estimate_number", "") or ""

    ws.cell(row=meta_row, column=1, value="案件名").font = _SUBTOTAL_FONT
    ws.merge_cells(start_row=meta_row, start_column=2, end_row=meta_row, end_column=4)
    ws.cell(row=meta_row, column=2, value=project_name).font = _BODY_FONT
    ws.cell(row=meta_row, column=5, value="顧客").font = _SUBTOTAL_FONT
    ws.merge_cells(start_row=meta_row, start_column=6, end_row=meta_row, end_column=8)
    ws.cell(row=meta_row, column=6, value=customer_name).font = _BODY_FONT

    ws.cell(row=meta_row + 1, column=1, value="見積番号").font = _SUBTOTAL_FONT
    ws.merge_cells(start_row=meta_row + 1, start_column=2, end_row=meta_row + 1, end_column=4)
    ws.cell(row=meta_row + 1, column=2, value=estimate_number).font = _BODY_FONT

    # ---- ヘッダー行 ----
    header_row = 6
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_HEADER
    ws.row_dimensions[header_row].height = 22

    # ---- 明細 ----
    row_idx = header_row + 1
    grand_total = 0
    item_seq = 0

    for section, items in sections_with_items:
        # 大項目ヘッダー (工事種別)
        section_name = getattr(section, "name", "") or ""
        section_code = getattr(section, "code", "") or ""
        section_label = f"【{section_code}】{section_name}" if section_code else section_name

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        sec_cell = ws.cell(row=row_idx, column=1, value=section_label)
        sec_cell.font = _SUBTOTAL_FONT
        sec_cell.fill = _SUBTOTAL_FILL
        sec_cell.alignment = _ALIGN_LEFT
        sec_cell.border = _BORDER_ALL
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

        section_subtotal = 0
        for item in items:
            item_seq += 1
            qty = _to_float(getattr(item, "quantity", 0))
            unit_price = _to_int(getattr(item, "sale_unit_price", 0))
            amount = _to_int(getattr(item, "sale_amount", 0))
            section_subtotal += amount

            values = [
                item_seq,
                getattr(item, "name", "") or "",
                getattr(item, "specification", "") or "",
                qty,
                getattr(item, "unit", "") or "",
                unit_price,
                amount,
                getattr(item, "note", "") or "",
            ]
            for i, v in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=i, value=v)
                cell.font = _BODY_FONT
                cell.border = _BORDER_ALL
                if i in (1, 5):
                    cell.alignment = _ALIGN_CENTER
                elif i in (4, 6, 7):
                    cell.alignment = _ALIGN_RIGHT
                else:
                    cell.alignment = _ALIGN_LEFT
            ws.cell(row=row_idx, column=4).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=6).number_format = '"¥"#,##0'
            ws.cell(row=row_idx, column=7).number_format = '"¥"#,##0'
            row_idx += 1

        # 種別小計
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        sub_label = ws.cell(row=row_idx, column=1, value=f"{section_name} 小計")
        sub_label.font = _SUBTOTAL_FONT
        sub_label.fill = _SUBTOTAL_FILL
        sub_label.alignment = _ALIGN_RIGHT
        sub_label.border = _BORDER_ALL

        sub_val = ws.cell(row=row_idx, column=7, value=section_subtotal)
        sub_val.font = _SUBTOTAL_FONT
        sub_val.fill = _SUBTOTAL_FILL
        sub_val.alignment = _ALIGN_RIGHT
        sub_val.border = _BORDER_ALL
        sub_val.number_format = '"¥"#,##0'

        empty = ws.cell(row=row_idx, column=8, value=None)
        empty.fill = _SUBTOTAL_FILL
        empty.border = _BORDER_ALL
        row_idx += 1
        grand_total += section_subtotal

    # ---- 合計 ----
    subtotal = _to_int(getattr(estimate, "subtotal", grand_total)) or grand_total
    tax_amount = _to_int(getattr(estimate, "tax_amount", 0))
    total = _to_int(getattr(estimate, "total", subtotal + tax_amount))
    tax_rate = _to_float(getattr(estimate, "tax_rate", 10.0))

    def _write_total_row(label: str, value: int, fill: PatternFill) -> int:
        nonlocal row_idx
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        l = ws.cell(row=row_idx, column=1, value=label)
        l.font = _SUBTOTAL_FONT
        l.fill = fill
        l.alignment = _ALIGN_RIGHT
        l.border = _BORDER_ALL

        v = ws.cell(row=row_idx, column=7, value=value)
        v.font = _SUBTOTAL_FONT
        v.fill = fill
        v.alignment = _ALIGN_RIGHT
        v.border = _BORDER_ALL
        v.number_format = '"¥"#,##0'

        e = ws.cell(row=row_idx, column=8, value=None)
        e.fill = fill
        e.border = _BORDER_ALL
        row_idx += 1
        return row_idx

    _write_total_row("小計 (税抜)", subtotal, _SUBTOTAL_FILL)
    _write_total_row(f"消費税 ({tax_rate:.0f}%)", tax_amount, _SUBTOTAL_FILL)
    _write_total_row("合計 (税込)", total, _TOTAL_FILL)

    # 印刷タイトル行 (ヘッダー行を各ページに繰り返す)
    ws.print_title_rows = f"{header_row}:{header_row}"


def write_estimate_detail(
    estimate,
    sections_with_items: Iterable[tuple[EstimateSection, list[EstimateItem]]],
) -> bytes:
    """見積内訳書を A4 印刷想定の xlsx として bytes 化する。

    Args:
        estimate: Estimate モデルインスタンス (project_name / customer_name / total 等)。
        sections_with_items: (EstimateSection, [EstimateItem,...]) のイテラブル。

    Returns:
        生成された xlsx の bytes。

    Raises:
        ValueError: estimate が None の場合。
    """
    if estimate is None:
        raise ValueError("estimate is required")

    sections_list: list[tuple[EstimateSection, list[EstimateItem]]] = list(sections_with_items)

    wb = Workbook()
    cover_ws = wb.active
    cover_ws.title = "中表紙"
    _write_cover_sheet(cover_ws, estimate)

    detail_ws = wb.create_sheet(title="内訳書")
    _write_detail_sheet(detail_ws, estimate, sections_list)

    # 内訳書を先頭タブに
    wb.move_sheet(detail_ws, offset=-1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_progress_sheet(
    ws: Worksheet,
    statement: ProgressStatement,
    rows_with_entries: Iterable[tuple[ProgressRow, list[ProgressEntry]]],
) -> None:
    """出来高調書シート本体。"""
    _setup_a4_landscape(ws)

    headers = [
        "No",
        "項目",
        "仕様",
        "単位",
        "契約数量",
        "契約単価",
        "契約金額",
        "当月数量",
        "当月金額",
        "累計数量",
        "累計金額",
        "進捗率 (%)",
    ]
    widths = [5, 22, 20, 8, 11, 13, 15, 11, 15, 11, 15, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- タイトル ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="出 来 高 調 書")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    # ---- メタ ----
    period_label = (
        getattr(statement, "period_label", None)
        or f"{getattr(statement, 'year', '')}年{getattr(statement, 'month', '')}月"
    )
    title = getattr(statement, "title", "") or ""
    version = getattr(statement, "version", "") or ""

    ws.cell(row=3, column=1, value="対象期").font = _SUBTOTAL_FONT
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
    ws.cell(row=3, column=2, value=period_label).font = _BODY_FONT

    ws.cell(row=3, column=5, value="タイトル").font = _SUBTOTAL_FONT
    ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=9)
    ws.cell(row=3, column=6, value=title).font = _BODY_FONT

    ws.cell(row=3, column=10, value="版").font = _SUBTOTAL_FONT
    ws.merge_cells(start_row=3, start_column=11, end_row=3, end_column=12)
    ws.cell(row=3, column=11, value=version).font = _BODY_FONT

    # ---- ヘッダー ----
    header_row = 5
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_HEADER
    ws.row_dimensions[header_row].height = 28

    # ---- 明細 ----
    row_idx = header_row + 1
    total_contract = 0
    total_current = 0
    total_cumulative = 0
    statement_year = getattr(statement, "year", None)
    statement_month = getattr(statement, "month", None)

    for seq, (row, entries) in enumerate(rows_with_entries, start=1):
        contract_qty = _to_float(getattr(row, "contract_qty", 0))
        contract_unit_price = _to_int(getattr(row, "contract_unit_price", 0))
        contract_amount = _to_int(getattr(row, "contract_amount", 0))
        cumulative_qty = _to_float(getattr(row, "cumulative_qty", 0))
        cumulative_amount = _to_int(getattr(row, "cumulative_amount", 0))

        # 当月エントリを抽出 (statement の year/month と一致するもの)
        current_qty = 0.0
        current_amount = 0
        for e in entries:
            if statement_year is not None and statement_month is not None:
                if getattr(e, "year", None) == statement_year and getattr(e, "month", None) == statement_month:
                    current_qty += _to_float(getattr(e, "progress_qty", 0))
                    current_amount += _to_int(getattr(e, "progress_amount", 0))
            else:
                current_qty += _to_float(getattr(e, "progress_qty", 0))
                current_amount += _to_int(getattr(e, "progress_amount", 0))

        progress_rate = 0.0
        if contract_amount > 0:
            progress_rate = (cumulative_amount / contract_amount) * 100.0

        values = [
            seq,
            getattr(row, "name", "") or "",
            getattr(row, "specification", "") or "",
            getattr(row, "contract_unit", "") or "",
            contract_qty,
            contract_unit_price,
            contract_amount,
            current_qty,
            current_amount,
            cumulative_qty,
            cumulative_amount,
            progress_rate,
        ]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=i, value=v)
            cell.font = _BODY_FONT
            cell.border = _BORDER_ALL
            if i in (1, 4):
                cell.alignment = _ALIGN_CENTER
            elif i in (2, 3):
                cell.alignment = _ALIGN_LEFT
            else:
                cell.alignment = _ALIGN_RIGHT

        # 数値フォーマット
        for col in (5, 8, 10):
            ws.cell(row=row_idx, column=col).number_format = "#,##0.00"
        for col in (6, 7, 9, 11):
            ws.cell(row=row_idx, column=col).number_format = '"¥"#,##0'
        ws.cell(row=row_idx, column=12).number_format = "0.0\"%\""

        total_contract += contract_amount
        total_current += current_amount
        total_cumulative += cumulative_amount
        row_idx += 1

    # ---- 合計行 ----
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    label_cell = ws.cell(row=row_idx, column=1, value="合 計")
    label_cell.font = _SUBTOTAL_FONT
    label_cell.fill = _TOTAL_FILL
    label_cell.alignment = _ALIGN_RIGHT
    label_cell.border = _BORDER_ALL

    contract_total_cell = ws.cell(row=row_idx, column=7, value=total_contract)
    current_total_cell = ws.cell(row=row_idx, column=9, value=total_current)
    cumulative_total_cell = ws.cell(row=row_idx, column=11, value=total_cumulative)
    overall_rate = (total_cumulative / total_contract * 100.0) if total_contract > 0 else 0.0
    rate_cell = ws.cell(row=row_idx, column=12, value=overall_rate)

    for col, cell in (
        (7, contract_total_cell),
        (9, current_total_cell),
        (11, cumulative_total_cell),
    ):
        cell.font = _SUBTOTAL_FONT
        cell.fill = _TOTAL_FILL
        cell.alignment = _ALIGN_RIGHT
        cell.border = _BORDER_ALL
        cell.number_format = '"¥"#,##0'

    rate_cell.font = _SUBTOTAL_FONT
    rate_cell.fill = _TOTAL_FILL
    rate_cell.alignment = _ALIGN_RIGHT
    rate_cell.border = _BORDER_ALL
    rate_cell.number_format = "0.0\"%\""

    # 空セルの塗り
    for col in (8, 10):
        c = ws.cell(row=row_idx, column=col, value=None)
        c.fill = _TOTAL_FILL
        c.border = _BORDER_ALL

    ws.print_title_rows = f"{header_row}:{header_row}"


def write_progress_statement(
    statement: ProgressStatement,
    rows_with_entries: Iterable[tuple[ProgressRow, list[ProgressEntry]]],
) -> bytes:
    """出来高調書を A4 印刷想定の xlsx として bytes 化する。

    Args:
        statement: ProgressStatement モデルインスタンス (year/month/title 等)。
        rows_with_entries: (ProgressRow, [ProgressEntry,...]) のイテラブル。

    Returns:
        生成された xlsx の bytes。

    Raises:
        ValueError: statement が None の場合。
    """
    if statement is None:
        raise ValueError("statement is required")

    wb = Workbook()
    ws = wb.active
    ws.title = "出来高調書"
    _write_progress_sheet(ws, statement, rows_with_entries)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
